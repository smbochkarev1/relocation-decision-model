#!/usr/bin/env python3
"""
relocation-decision-model — weighted multi-criteria scoring engine.

Reads a configuration (criteria.yaml + scores.csv, optional knockouts.csv),
applies knockout filters, computes a weighted score for every surviving
candidate, ranks them, runs a sensitivity analysis over alternative weight
presets, and writes an Excel workbook + a Markdown report.

The engine is generic: it has no knowledge of countries, criteria, or the
author's situation. Everything lives in the config files. Score every
criterion on a 1-10 scale (10 = best for you). Weights are expressed in
percent and should sum to 100 (the engine normalizes by their sum regardless).

Usage:
    python score.py                         # uses ./config
    python score.py --config examples/eu-relocation-couple
    python score.py --config <dir> --out <dir> --top 5

Dependencies: pyyaml, openpyxl (see requirements.txt).
"""
from __future__ import annotations

import argparse
import csv
import os
import sys
from dataclasses import dataclass, field

try:
    import yaml
except ImportError:
    sys.exit("Missing dependency 'pyyaml'. Install with: pip install -r requirements.txt")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
except ImportError:
    sys.exit("Missing dependency 'openpyxl'. Install with: pip install -r requirements.txt")


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------
@dataclass
class Criterion:
    cid: str
    name: str
    weight: float
    domain_id: str
    domain_name: str
    source: str = ""


@dataclass
class Config:
    profile: str
    title: str
    leader_note: str
    disclaimers: list
    criteria: list  # list[Criterion], in declared order
    knockouts: list  # list[dict(id, label)]
    sensitivity: list  # list[dict(name, overrides)]


TRUEISH = {"pass", "yes", "y", "true", "1", "ok", "t"}
FALSEISH = {"fail", "no", "n", "false", "0", "x", "f"}


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------
def load_config(path: str) -> Config:
    with open(path, "r", encoding="utf-8") as fh:
        raw = yaml.safe_load(fh)

    criteria: list[Criterion] = []
    for dom in raw.get("domains", []):
        did = str(dom["id"])
        dname = dom.get("name", did)
        for c in dom.get("criteria", []):
            criteria.append(
                Criterion(
                    cid=str(c["id"]),
                    name=c["name"],
                    weight=float(c["weight"]),
                    domain_id=did,
                    domain_name=dname,
                    source=c.get("source", ""),
                )
            )
    if not criteria:
        sys.exit("Config error: no criteria found under 'domains'.")

    report = raw.get("report", {}) or {}
    return Config(
        profile=str(raw.get("profile", "")).strip(),
        title=report.get("title", "Relocation scoring"),
        leader_note=report.get("leader_note", ""),
        disclaimers=report.get("disclaimers", []) or [],
        criteria=criteria,
        knockouts=raw.get("knockouts", []) or [],
        sensitivity=raw.get("sensitivity", []) or [{"name": "Baseline", "overrides": {}}],
    )


def load_scores(path: str, criteria: list) -> tuple:
    """Return (scores: {country: {cid: float}}, regions: {country: str})."""
    cids = {c.cid for c in criteria}
    scores: dict = {}
    regions: dict = {}
    with open(path, "r", encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        header = reader.fieldnames or []
        missing = cids - set(header)
        if missing:
            sys.exit(
                "scores.csv is missing columns for criteria: "
                + ", ".join(sorted(missing))
            )
        for row in reader:
            country = (row.get("country") or "").strip()
            if not country:
                continue
            regions[country] = (row.get("region") or "").strip()
            vals = {}
            for cid in cids:
                cell = (row.get(cid) or "").strip()
                if cell == "":
                    sys.exit(f"scores.csv: empty score for {country} / {cid}.")
                try:
                    v = float(cell)
                except ValueError:
                    sys.exit(f"scores.csv: non-numeric score '{cell}' for {country} / {cid}.")
                if not (1.0 <= v <= 10.0):
                    sys.exit(f"scores.csv: score {v} for {country} / {cid} out of range [1, 10].")
                vals[cid] = v
            scores[country] = vals
    if not scores:
        sys.exit("scores.csv contains no data rows.")
    return scores, regions


def load_knockouts(path: str, knockout_defs: list) -> dict:
    """Return {country: {"failed": [labels], "reason": str, "region": str}}."""
    result: dict = {}
    if not os.path.exists(path):
        return result
    label_of = {str(k["id"]): k.get("label", str(k["id"])) for k in knockout_defs}
    with open(path, "r", encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        header = reader.fieldnames or []
        ko_cols = [h for h in header if h in label_of]
        for row in reader:
            country = (row.get("country") or "").strip()
            if not country:
                continue
            failed = []
            for col in ko_cols:
                cell = (row.get(col) or "").strip().lower()
                if cell in FALSEISH:
                    failed.append(label_of[col])
                elif cell and cell not in TRUEISH:
                    sys.exit(f"knockouts.csv: unrecognized value '{cell}' for {country} / {col}.")
            result[country] = {
                "failed": failed,
                "reason": (row.get("reason") or "").strip(),
                "region": (row.get("region") or "").strip(),
            }
    return result


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------
def weighted_score(vals: dict, weights: dict, cids: list) -> float:
    total_w = sum(weights[c] for c in cids)
    if total_w == 0:
        return 0.0
    return sum(vals[c] * weights[c] for c in cids) / total_w


def rank_scenario(scores: dict, weights: dict, cids: list) -> tuple:
    """Return (score_map, rank_map, ordered[(country, score)])."""
    score_map = {c: round(weighted_score(v, weights, cids), 2) for c, v in scores.items()}
    ordered = sorted(score_map.items(), key=lambda kv: (-kv[1], kv[0]))
    rank_map = {name: i + 1 for i, (name, _) in enumerate(ordered)}
    return score_map, rank_map, ordered


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
FONT = "Calibri"
HDR = PatternFill("solid", fgColor="1F3864")
DOM = PatternFill("solid", fgColor="2E5496")
WT = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WHITE_B = Font(name=FONT, bold=True, color="FFFFFF", size=9)
BLK = Font(name=FONT, size=9)
BLK_B = Font(name=FONT, size=9, bold=True)
BLUE = Font(name=FONT, size=9, color="0000FF", bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")


def write_xlsx(path, cfg, scores, regions, base_rank, base_scores, base_order,
               sensitivity_results, knocked_out):
    cids = [c.cid for c in cfg.criteria]
    weights = {c.cid: c.weight for c in cfg.criteria}
    wb = Workbook()

    # ---- Sheet 1: Scoring ----
    ws = wb.active
    ws.title = "Scoring"
    col0 = 3
    n = len(cids)
    total_col = col0 + n
    rank_col = total_col + 1
    first_L, last_L = get_column_letter(col0), get_column_letter(col0 + n - 1)
    total_L = get_column_letter(total_col)

    ws.cell(1, 1, "Country").font = WHITE_B
    ws.cell(1, 1).fill = HDR
    ws.cell(1, 2, "Region").font = WHITE_B
    ws.cell(1, 2).fill = HDR
    ws.cell(3, 1, "WEIGHT, % →").font = BLK_B
    for j, c in enumerate(cfg.criteria):
        col = col0 + j
        L = get_column_letter(col)
        dc = ws.cell(1, col, c.domain_name)
        dc.font = WHITE_B; dc.fill = DOM; dc.alignment = CENTER
        pc = ws.cell(2, col, f"{c.cid}. {c.name}")
        pc.font = WHITE_B; pc.fill = HDR; pc.alignment = CENTER
        wc = ws.cell(3, col, c.weight)
        wc.font = BLUE; wc.fill = WT; wc.alignment = CENTER; wc.border = BORDER
        ws.column_dimensions[L].width = 7
    ws.cell(1, total_col, "TOTAL").font = WHITE_B
    ws.cell(1, total_col).fill = HDR; ws.cell(1, total_col).alignment = CENTER
    ws.cell(2, total_col, "0–10").font = WHITE_B
    ws.cell(2, total_col).fill = HDR; ws.cell(2, total_col).alignment = CENTER
    ws.cell(3, total_col, f"=SUM({first_L}3:{last_L}3)").font = BLK_B
    ws.cell(1, rank_col, "RANK").font = WHITE_B
    ws.cell(1, rank_col).fill = HDR; ws.cell(1, rank_col).alignment = CENTER

    ordered_countries = [name for name, _ in base_order]
    r0 = 4
    for i, name in enumerate(ordered_countries):
        r = r0 + i
        ws.cell(r, 1, name).font = BLK_B
        ws.cell(r, 1).alignment = LEFT; ws.cell(r, 1).border = BORDER
        ws.cell(r, 2, regions.get(name, "")).font = BLK
        ws.cell(r, 2).alignment = LEFT; ws.cell(r, 2).border = BORDER
        for j, cid in enumerate(cids):
            v = ws.cell(r, col0 + j, scores[name][cid])
            v.font = BLK; v.alignment = CENTER; v.border = BORDER; v.number_format = "0.0"
        it = ws.cell(
            r, total_col,
            f"=SUMPRODUCT({first_L}{r}:{last_L}{r},${first_L}$3:${last_L}$3)"
            f"/SUM(${first_L}$3:${last_L}$3)",
        )
        it.font = BLK_B; it.alignment = CENTER; it.border = BORDER; it.number_format = "0.00"
        rk = ws.cell(r, rank_col, base_rank[name])
        rk.font = BLK_B; rk.alignment = CENTER; rk.border = BORDER
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions[total_L].width = 8
    ws.freeze_panes = "C4"
    last_r = r0 + len(ordered_countries) - 1
    ws.conditional_formatting.add(
        f"{first_L}{r0}:{last_L}{last_r}",
        ColorScaleRule(start_type="num", start_value=1, start_color="F8696B",
                       mid_type="num", mid_value=5.5, mid_color="FFEB84",
                       end_type="num", end_value=10, end_color="63BE7B"))
    ws.conditional_formatting.add(
        f"{total_L}{r0}:{total_L}{last_r}",
        ColorScaleRule(start_type="min", start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))

    # ---- Sheet 2: Sensitivity ----
    ws2 = wb.create_sheet("Sensitivity")
    ws2.cell(1, 1, "Weight scenarios — score and rank per country").font = Font(
        name=FONT, bold=True, size=11)
    hr = 3
    ws2.cell(hr, 1, "Country").font = WHITE_B; ws2.cell(hr, 1).fill = HDR
    ws2.column_dimensions["A"].width = 20
    col = 2
    scen_cols = {}
    for scen in cfg.sensitivity:
        nm = scen["name"]
        ws2.cell(hr, col, nm + "\nscore").font = WHITE_B
        ws2.cell(hr, col).fill = HDR; ws2.cell(hr, col).alignment = CENTER
        ws2.cell(hr, col + 1, "rank").font = WHITE_B
        ws2.cell(hr, col + 1).fill = HDR; ws2.cell(hr, col + 1).alignment = CENTER
        ws2.column_dimensions[get_column_letter(col)].width = 12
        ws2.column_dimensions[get_column_letter(col + 1)].width = 6
        scen_cols[nm] = col
        col += 2
    for i, name in enumerate(ordered_countries):
        r = hr + 1 + i
        ws2.cell(r, 1, name).font = BLK
        for scen in cfg.sensitivity:
            nm = scen["name"]
            sm, rm = sensitivity_results[nm]
            c = scen_cols[nm]
            vc = ws2.cell(r, c, sm[name]); vc.number_format = "0.00"
            vc.font = BLK; vc.alignment = CENTER
            rc = ws2.cell(r, c + 1, rm[name]); rc.font = BLK_B; rc.alignment = CENTER
    last2 = hr + len(ordered_countries)
    for scen in cfg.sensitivity:
        cL = get_column_letter(scen_cols[scen["name"]])
        ws2.conditional_formatting.add(
            f"{cL}{hr+1}:{cL}{last2}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"))
    ws2.freeze_panes = "B4"

    # ---- Sheet 3: Knockouts ----
    if knocked_out:
        ws3 = wb.create_sheet("Knockouts")
        ws3.cell(1, 1, "Candidates excluded before scoring").font = Font(
            name=FONT, bold=True, size=11)
        heads = ["Country", "Region", "Failed filter(s)", "Note"]
        for j, h in enumerate(heads, 1):
            c = ws3.cell(3, j, h); c.font = WHITE_B; c.fill = HDR; c.alignment = CENTER
        for i, ko in enumerate(knocked_out):
            r = 4 + i
            ws3.cell(r, 1, ko["country"]).font = BLK_B
            ws3.cell(r, 2, ko.get("region", "")).font = BLK
            ws3.cell(r, 3, "; ".join(ko["failed"])).font = BLK
            ws3.cell(r, 4, ko.get("reason", "")).font = BLK
            for j in range(1, 5):
                ws3.cell(r, j).alignment = Alignment(wrap_text=True, vertical="top")
                ws3.cell(r, j).border = BORDER
        for L, w in zip("ABCD", (20, 20, 34, 40)):
            ws3.column_dimensions[L].width = w

    # ---- Sheet 4: Anchors / methodology ----
    ws4 = wb.create_sheet("Anchors")
    ws4.column_dimensions["A"].width = 120
    lines = [(cfg.title, True), ("", False)]
    if cfg.profile:
        lines.append(("PROFILE: " + cfg.profile.replace("\n", " "), False))
        lines.append(("", False))
    lines.append(("HOW TO USE: edit the yellow weight row on the 'Scoring' sheet — TOTAL and RANK recompute automatically.", False))
    lines.append(("", False))
    lines.append(("CRITERIA & ANCHORS (1–3 weak / 4–6 acceptable / 7–10 strong):", True))
    _anchor_lines_written = _write_anchor_lines(ws4, cfg, lines)
    for i, (txt, bold) in enumerate(_anchor_lines_written, 1):
        c = ws4.cell(i, 1, txt)
        c.font = Font(name=FONT, bold=bold, size=11 if (bold and i == 1) else 9)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)


def _write_anchor_lines(ws, cfg, lines):
    """Kept simple: anchors are surfaced in report.md; here just list weights."""
    for c in cfg.criteria:
        lines.append((f"{c.cid}. {c.name}  (domain: {c.domain_name}, weight {c.weight})", False))
        if c.source:
            lines.append((f"     source: {c.source}", False))
    return lines


# ---------------------------------------------------------------------------
# Markdown report
# ---------------------------------------------------------------------------
def write_report(path, cfg, base_scores, base_rank, base_order,
                 sensitivity_results, regions, knocked_out, top_n):
    L = []
    w = L.append
    w(f"# {cfg.title}\n")
    if cfg.profile:
        w(f"*{cfg.profile.strip()}*\n")
    w("*Generated by `score.py`. Every criterion scored 1–10 (10 = best). "
      f"Weighted total = Σ(score × weight) / Σ(weight). {len(base_order)} candidates scored.*\n")
    w("\n---\n")

    # Headline
    (leader, l_score) = base_order[0]
    w("\n## Bottom line\n")
    if cfg.leader_note:
        w("\n" + cfg.leader_note.strip() + "\n")
    else:
        gap = ""
        if len(base_order) > 1:
            second, s_score = base_order[1]
            gap = (f" **{leader}** leads with **{l_score}**; "
                   f"**{second}** is second at **{s_score}** "
                   f"(gap {round(l_score - s_score, 2)}).")
        w(f"\n{gap}\n")

    # Ranking table
    w("\n## Ranking\n")
    w("\n| # | Country | Region | Score |")
    w("\n|---|---------|--------|-------|")
    for name, sc in base_order:
        w(f"\n| {base_rank[name]} | {name} | {regions.get(name, '')} | {sc:.2f} |")
    w("\n")

    # Knockouts
    if knocked_out:
        w("\n## Knocked out (excluded before scoring)\n")
        w("\n| Country | Failed filter(s) | Note |")
        w("\n|---------|------------------|------|")
        for ko in knocked_out:
            w(f"\n| {ko['country']} | {'; '.join(ko['failed'])} | {ko.get('reason','')} |")
        w("\n")

    # Sensitivity
    w("\n## Sensitivity: how stable is the leader?\n")
    w(f"\n| Scenario | Top {top_n} |")
    w("\n|----------|--------|")
    for scen in cfg.sensitivity:
        nm = scen["name"]
        sm, rm = sensitivity_results[nm]
        top = [name for name, _ in sorted(sm.items(), key=lambda kv: (-kv[1], kv[0]))[:top_n]]
        w(f"\n| {nm} | {', '.join(top)} |")
    w("\n")

    # Data sources
    srcs = [(c.cid, c.name, c.source) for c in cfg.criteria if c.source]
    if srcs:
        w("\n## Data sources\n")
        w("\n| Criterion | Source |")
        w("\n|-----------|--------|")
        for cid, name, src in srcs:
            w(f"\n| {cid}. {name} | {src} |")
        w("\n")

    # Disclaimers
    if cfg.disclaimers:
        w("\n## Caveats\n")
        for d in cfg.disclaimers:
            w(f"\n- {d}")
        w("\n")

    with open(path, "w", encoding="utf-8") as fh:
        fh.write("".join(L))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    ap = argparse.ArgumentParser(description="Weighted multi-criteria relocation scoring.")
    ap.add_argument("--config", default="config",
                    help="Config directory (criteria.yaml, scores.csv, [knockouts.csv]).")
    ap.add_argument("--out", default=None,
                    help="Output directory (default: <config>/output).")
    ap.add_argument("--top", type=int, default=5, help="Top-N shown in sensitivity/report.")
    args = ap.parse_args()

    cfg_dir = args.config
    out_dir = args.out or os.path.join(cfg_dir, "output")
    os.makedirs(out_dir, exist_ok=True)

    cfg = load_config(os.path.join(cfg_dir, "criteria.yaml"))
    cids = [c.cid for c in cfg.criteria]
    weights = {c.cid: c.weight for c in cfg.criteria}
    wsum = sum(weights.values())
    if abs(wsum - 100) > 1e-6:
        print(f"WARNING: weights sum to {wsum}, not 100. Engine normalizes by the sum anyway.")

    scores, regions = load_scores(os.path.join(cfg_dir, "scores.csv"), cfg.criteria)
    ko_data = load_knockouts(os.path.join(cfg_dir, "knockouts.csv"), cfg.knockouts)

    # Apply knockouts across the candidate universe (scores.csv ∪ knockouts.csv).
    universe = set(scores) | set(ko_data)
    knocked_out = []
    survivors = []
    for country in sorted(universe):
        info = ko_data.get(country, {})
        failed = info.get("failed", [])
        if failed:
            reg = info.get("region") or regions.get(country, "")
            knocked_out.append({"country": country, "failed": failed,
                                "reason": info.get("reason", ""), "region": reg})
        else:
            if country not in scores:
                sys.exit(f"'{country}' passes knockouts but has no row in scores.csv.")
            survivors.append(country)
    if not survivors:
        sys.exit("No candidates survived the knockout filters.")
    scores = {c: scores[c] for c in survivors}

    # Baseline ranking
    base_scores, base_rank, base_order = rank_scenario(scores, weights, cids)

    # Sensitivity
    sensitivity_results = {}
    for scen in cfg.sensitivity:
        w2 = dict(weights)
        for k, v in (scen.get("overrides") or {}).items():
            if k not in w2:
                sys.exit(f"Sensitivity preset '{scen['name']}' overrides unknown criterion '{k}'.")
            w2[k] = float(v)
        sm, rm, _ = rank_scenario(scores, w2, cids)
        sensitivity_results[scen["name"]] = (sm, rm)

    xlsx_path = os.path.join(out_dir, "ranking.xlsx")
    md_path = os.path.join(out_dir, "report.md")
    write_xlsx(xlsx_path, cfg, scores, regions, base_rank, base_scores, base_order,
               sensitivity_results, knocked_out)
    write_report(md_path, cfg, base_scores, base_rank, base_order,
                 sensitivity_results, regions, knocked_out, args.top)

    print(f"Scored {len(survivors)} candidates | knocked out {len(knocked_out)} | "
          f"weights sum {wsum:g}")
    print(f"Wrote {xlsx_path}")
    print(f"Wrote {md_path}")
    print(f"\nTop {args.top}:")
    for name, sc in base_order[:args.top]:
        print(f"  {base_rank[name]:>2}. {name:<20} {sc:.2f}")


if __name__ == "__main__":
    main()
